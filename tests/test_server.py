"""任务存储、修订工具与 HTTP/飞书适配。"""

from __future__ import annotations

import http.client
import json
import os
import threading
import time
from http.server import ThreadingHTTPServer
from pathlib import Path

import pytest

from qagent.agent.llm import MockLLM
from qagent.parsing import parse_cases
from qagent.config import public_llm_settings, update_local_llm
from qagent.server.app import create_handler
from qagent.server.chat import apply_actions, run_chat
from qagent.server.feishu import handle_feishu_event
from qagent.server.jobs import JobStore
from qagent.server.service import QAgentService
from qagent.server.tools import delete_cases, patch_plan, upsert_cases, validate_and_export

FIXTURES = Path(__file__).parent / "fixtures"


def _dummy_api_key() -> str:
    """返回占位 API key，测试用，非真实凭据。"""
    return os.environ.get("QAGENT_TEST_API_KEY", "dummy-placeholder-for-tests")


def _http(port: int, method: str, path: str, body: bytes | None = None,
          headers: dict | None = None) -> tuple[int, bytes]:
    """通过 http.client 直连本地测试服务器，避免 SSRF 风险。"""
    conn = http.client.HTTPConnection("127.0.0.1", port, timeout=5)
    try:
        hdrs: dict[str, str] = {"Host": f"127.0.0.1:{port}"}
        if headers:
            hdrs.update(headers)
        if body is not None:
            hdrs["Content-Length"] = str(len(body))
        conn.request(method, path, body=body, headers=hdrs)
        resp = conn.getresponse()
        return resp.status, resp.read()
    finally:
        conn.close()


def _http_json(port: int, method: str, path: str, body: bytes | None = None,
               headers: dict | None = None) -> dict:
    """发起请求并解析 JSON 响应。"""
    _status, data = _http(port, method, path, body, headers)
    return json.loads(data)


def _seed_output(store: JobStore, job_id: str) -> None:
    out = store.output_dir(job_id)
    for name in ("test-plan.md", "risk.md", "coverage-matrix.md", "testcases-valid.md"):
        src = FIXTURES / name
        dest = out / ("testcases.md" if name == "testcases-valid.md" else name)
        dest.write_text(src.read_text(encoding="utf-8"), encoding="utf-8")
    treq = FIXTURES / "test-requirements-generated.md"
    if treq.is_file():
        (out / "test-requirements.md").write_text(treq.read_text(encoding="utf-8"), encoding="utf-8")
    review = FIXTURES / "qa-review.md"
    if review.is_file():
        (out / "qa-review.md").write_text(review.read_text(encoding="utf-8"), encoding="utf-8")


def test_job_store_isolation(tmp_path):
    store = JobStore(tmp_path / "jobs")
    a = store.create(owner="alice", title="A")
    b = store.create(owner="bob", title="B")
    store.save_upload(a.id, "prd.md", b"# A\n")
    store.save_upload(b.id, "prd.md", b"# B\n")
    assert (store.input_dir(a.id) / "prd.md").read_text() == "# A\n"
    assert (store.input_dir(b.id) / "prd.md").read_text() == "# B\n"
    assert len(store.list_jobs()) == 2
    assert store.load(a.id).owner == "alice"


def test_delete_job_removes_files(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create(title="gone")
    store.save_upload(job.id, "req.md", b"# x\n")
    store.delete(job.id)
    with pytest.raises(FileNotFoundError):
        store.load(job.id)
    assert not (tmp_path / "jobs" / job.id).exists()


def test_delete_running_job_rejected(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    meta = store.load(job.id)
    meta.status = "running"
    store.save_meta(meta)
    with pytest.raises(RuntimeError, match="执行中"):
        store.delete(job.id)
    assert store.load(job.id).id == job.id


def test_start_run_clears_previous_error_and_logs(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    job = store.create()
    store.save_upload(job.id, "req.md", b"# hello\n")
    meta = store.load(job.id)
    meta.status = "failed"
    meta.error = ["读取 PDF 需要安装 pypdf"]
    meta.logs = ["[19:21:31] ERROR: 读取 PDF 需要安装 pypdf (OCR-PRD.pdf)"]
    store.save_meta(meta)
    public = service.start_run(job.id, "requirements")
    assert public["status"] == "running"
    assert public["error"] is None
    assert public["logs"] == []


def test_start_run_allows_parallel_jobs(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}), max_pipeline=2)
    a = store.create(title="A")
    b = store.create(title="B")
    store.save_upload(a.id, "a.md", b"# A\n")
    store.save_upload(b.id, "b.md", b"# B\n")
    ra = service.start_run(a.id)
    rb = service.start_run(b.id)
    assert ra["status"] == "running"
    assert rb["status"] == "running"
    assert ra["id"] != rb["id"]


def test_http_delete_job(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    httpd = ThreadingHTTPServer(("127.0.0.1", 0), create_handler(service))
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    port = httpd.server_address[1]
    try:
        job = service.create_job("t", [("req.md", b"# hello\n")])
        status, data = _http(port, "DELETE", f"/api/jobs/{job['id']}")
        body = json.loads(data)
        assert body["ok"] is True
        status, _ = _http(port, "GET", f"/api/jobs/{job['id']}")
        assert status == 404
        listed = _http_json(port, "GET", "/api/jobs")
        assert listed["jobs"] == []
    finally:
        httpd.shutdown()


def test_patch_upsert_delete_and_validate(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    _seed_output(store, job.id)
    patch_plan(store, job.id, add=[{"id": "R9", "text": "补一条可验证规则"}])
    text = (store.output_dir(job.id) / "test-plan.md").read_text(encoding="utf-8")
    assert "R9: 补一条可验证规则" in text
    patch_plan(store, job.id, edit={"R9": ""})  # 不把额外 R 带进校验
    (store.output_dir(job.id) / "test-plan.md").write_text(
        (FIXTURES / "test-plan.md").read_text(encoding="utf-8"), encoding="utf-8",
    )
    upsert_cases(store, job.id, [{
        "id": "TC-REG-009",
        "title": "SC-001 额外说明",
        "priority": "P1",
        "type": "功能",
        "preconditions": [],
        "steps": ["打开注册页"],
        "expected": "页面可见",
        "design_method": "场景法",
        "requirement_ref": "R1",
    }])
    delete_cases(store, job.id, ["TC-REG-009"])
    result = validate_and_export(store, job.id, fill_gaps=True)
    assert result["ok"], result.get("errors")
    assert (store.output_dir(job.id) / "testcases.xlsx").is_file()
    assert (store.output_dir(job.id) / "test-requirements.drawio").is_file()
    assert not (store.output_dir(job.id) / "test-plan.drawio").exists()


def test_upsert_fills_missing_requirement_ref(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    _seed_output(store, job.id)
    upsert_cases(store, job.id, [{
        "id": "TC-PERF-001",
        "title": "简单图片识别性能",
        "priority": "P1",
        "type": "性能",
        "preconditions": [],
        "steps": ["上传100字图片并计时"],
        "expected": "4秒内返回结果",
        "design_method": "场景法",
    }])
    cases = parse_cases(store.output_dir(job.id) / "testcases.md")
    perf = next(case for case in cases if case["id"] == "TC-PERF-001")
    assert perf["requirement_ref"]
    assert perf["type"] == "功能"
    from openpyxl import load_workbook
    ws = load_workbook(store.output_dir(job.id) / "testcases.xlsx").active
    xlsx_ids = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert "TC-PERF-001" in xlsx_ids
    result = validate_and_export(store, job.id, fill_gaps=True)
    assert result["ok"], result


def test_upsert_without_validate_still_updates_xlsx(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    _seed_output(store, job.id)
    notes, validated, _ = apply_actions(store, job.id, [{
        "op": "upsert_cases",
        "cases": [{
            "id": "TC-PERF-002",
            "title": "复杂文档识别耗时",
            "priority": "P1",
            "type": "功能",
            "preconditions": [],
            "steps": ["上传复杂图并计时"],
            "expected": "15秒内返回",
            "design_method": "场景法",
            "requirement_ref": "R1",
        }],
    }])
    assert validated is None
    assert "已合并" in notes[0]
    from openpyxl import load_workbook
    ws = load_workbook(store.output_dir(job.id) / "testcases.xlsx").active
    xlsx_ids = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert "TC-PERF-002" in xlsx_ids


def test_apply_actions_rollback_on_bad_delete(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    _seed_output(store, job.id)
    before = (store.output_dir(job.id) / "testcases.md").read_text(encoding="utf-8")

    class DelLLM:
        def complete(self, system, user):
            return json.dumps({
                "reply": "已删除",
                "actions": [
                    {"op": "delete_cases", "ids": ["TC-REG-001", "TC-REG-002", "TC-REG-003"]},
                    {"op": "validate_and_export", "fill_gaps": False},
                ],
            })

    result = run_chat(store, job.id, "删掉所有用例", DelLLM())
    assert not result["ok"]
    after = (store.output_dir(job.id) / "testcases.md").read_text(encoding="utf-8")
    assert after == before


def test_run_chat_json_actions(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    _seed_output(store, job.id)
    class JsonLLM:
        def complete(self, system, user):
            return json.dumps({
                "reply": "已补用例",
                "actions": [
                    {
                        "op": "upsert_cases",
                        "cases": [{
                            "id": "TC-REG-008",
                            "title": "SC-001 对话补的用例",
                            "priority": "P1",
                            "type": "功能",
                            "preconditions": [],
                            "steps": ["打开注册"],
                            "expected": "可打开",
                            "design_method": "场景法",
                            "requirement_ref": "R1",
                        }],
                    },
                    {"op": "validate_and_export", "fill_gaps": True},
                ],
            })
    result = run_chat(store, job.id, "给 R1 补一条用例", JsonLLM())
    assert result["ok"], result
    cases = (store.output_dir(job.id) / "testcases.md").read_text(encoding="utf-8")
    assert "TC-REG-008" in cases
    from openpyxl import load_workbook
    ws = load_workbook(store.output_dir(job.id) / "testcases.xlsx").active
    xlsx_ids = [ws.cell(row, 1).value for row in range(2, ws.max_row + 1)]
    assert "TC-REG-008" in xlsx_ids
    chat = store.load_chat(job.id)
    assert [m["role"] for m in chat] == ["user", "assistant"]


def test_start_chat_returns_before_llm_finishes(tmp_path):
    store = JobStore(tmp_path / "jobs")
    started = threading.Event()
    release = threading.Event()

    class SlowLLM:
        def complete(self, system, user):
            started.set()
            assert release.wait(timeout=2)
            return json.dumps({"reply": "已收到", "actions": []})

    service = QAgentService(store, llm_factory=lambda: SlowLLM(), max_pipeline=2)
    job = store.create()
    _seed_output(store, job.id)
    public = service.start_chat(job.id, "补一条边界用例")
    assert public["status"] == "revising"
    assert public["chat"][-1]["role"] == "user"
    assert public["chat"][-1]["content"] == "补一条边界用例"
    assert started.wait(timeout=2)
    release.set()
    got = None
    for _ in range(80):
        got = service.get_job(job.id)
        if got["status"] == "ready":
            break
        time.sleep(0.05)
    assert got is not None
    assert got["status"] == "ready"
    assert [m["role"] for m in got["chat"][-2:]] == ["user", "assistant"]
    assert "已收到" in got["chat"][-1]["content"]


def test_http_multipart_upload_creates_job(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    httpd = ThreadingHTTPServer(("127.0.0.1", 0), create_handler(service))
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    port = httpd.server_address[1]
    try:
        boundary = "----QAgentTestBoundary"
        payload = (
            f"--{boundary}\r\n"
            'Content-Disposition: form-data; name="files"; filename="req.md"\r\n'
            "Content-Type: text/markdown\r\n"
            "\r\n"
            "# 登录\n用户可登录。\n"
            f"\r\n--{boundary}--\r\n"
        ).encode("utf-8")
        job = _http_json(port, "POST", "/api/jobs", body=payload, headers={
            "Content-Type": f"multipart/form-data; boundary={boundary}",
            "X-User": "alice",
        })
        assert job["id"]
        assert job.get("awaiting_scope") is True
        got = _http_json(port, "GET", f"/api/jobs/{job['id']}")
        assert "req.md" in got["inputs"]
    finally:
        httpd.shutdown()


def test_http_jobs_and_health(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    httpd = ThreadingHTTPServer(("127.0.0.1", 0), create_handler(service))
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    port = httpd.server_address[1]
    try:
        health = _http_json(port, "GET", "/health")
        assert health["ok"]
        listed = _http_json(port, "GET", "/api/jobs")
        assert listed["jobs"] == []
        job = service.create_job("t", [("req.md", b"# hello\n")])
        got = _http_json(port, "GET", f"/api/jobs/{job['id']}")
        assert got["id"] == job["id"]
        assert "req.md" in got["inputs"]
    finally:
        httpd.shutdown()


def test_feishu_url_verification(tmp_path):
    service = QAgentService(JobStore(tmp_path / "jobs"), llm_factory=lambda: MockLLM({}))
    out = handle_feishu_event(service, {"type": "url_verification", "challenge": "abc"})
    assert out["challenge"] == "abc"


def test_create_job_asks_scope_without_test_requirements(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    job = service.create_job("t", [("prd.md", "# 产品\n登录\n".encode("utf-8"))])
    assert job["awaiting_scope"] is True
    assert any("测试范围" in m["content"] for m in job["chat"])


def test_create_job_skips_scope_when_test_requirements_uploaded(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    job = service.create_job("t", [("测试需求.md", "# 测试需求\n全量\n".encode("utf-8"))])
    assert job["awaiting_scope"] is False


def test_scope_confirm_starts_run(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    job = service.create_job("t", [("prd.md", "# 产品\n登录\n".encode("utf-8"))])
    result = service.chat(job["id"], "可以")
    assert result["rerun"] == "requirements"
    assert (store.input_dir(job["id"]) / "测试需求.md").is_file()
    assert store.load(job["id"]).awaiting_scope is False
    assert store.load(job["id"]).status in {"running", "ready", "failed"}


def test_start_run_from_testcases_requires_matrix(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    job = store.create()
    store.save_upload(job.id, "req.md", b"# hello\n")
    with pytest.raises(RuntimeError, match="覆盖矩阵"):
        service.start_run(job.id, "testcases")
    _seed_output(store, job.id)
    public = service.start_run(job.id, "testcases")
    assert public["status"] == "running"


def test_cancel_job_marks_cancelled(tmp_path):
    store = JobStore(tmp_path / "jobs")
    started = threading.Event()
    release = threading.Event()

    class SlowLLM:
        def complete(self, system, user):
            started.set()
            release.wait(timeout=3)
            return "ok"

    service = QAgentService(store, llm_factory=lambda: SlowLLM(), max_pipeline=2)
    job = store.create()
    store.save_upload(job.id, "req.md", b"# hello\n")
    service.start_run(job.id, "requirements")
    assert started.wait(timeout=2)
    public = service.cancel_job(job.id)
    assert public["cancel_requested"] is True
    release.set()
    got = None
    for _ in range(80):
        got = service.get_job(job.id)
        if got["status"] == "cancelled":
            break
        time.sleep(0.05)
    assert got is not None
    assert got["status"] == "cancelled"
    service.start_run(job.id, "requirements")
    assert store.load(job.id).status == "running"
    release.set()


def test_feishu_rebind_creates_new_job_keeps_old(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    first = service.create_job("u", [("a.md", b"# A\n")])
    second = service.create_job("u", [("b.md", b"# B\n")])
    store.bind_feishu(first["id"], "oc_chat", "u")
    store.bind_feishu(second["id"], "oc_chat", "u")
    assert store.job_for_feishu_chat("oc_chat") == second["id"]
    assert store.load(first["id"]).id == first["id"]


def test_read_artifact_performance_synonyms(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    _seed_output(store, job.id)
    cases = store.output_dir(job.id) / "testcases.md"
    cases.write_text(cases.read_text(encoding="utf-8") + "\n总耗时 T ≤ 4秒\n", encoding="utf-8")
    from qagent.server.tools import read_artifact
    found = read_artifact(store, job.id, "cases", "性能")
    assert "耗时" in found


def test_chat_keeps_good_case_drops_bad_on_validate(tmp_path):
    store = JobStore(tmp_path / "jobs")
    job = store.create()
    _seed_output(store, job.id)
    before_ids = {c["id"] for c in parse_cases(store.output_dir(job.id) / "testcases.md")}

    class MixedLLM:
        def complete(self, system, user):
            return json.dumps({
                "reply": "已补",
                "actions": [
                    {
                        "op": "upsert_cases",
                        "cases": [
                            {
                                "id": "TC-REG-008",
                                "title": "SC-001 好用例",
                                "priority": "P1",
                                "type": "功能",
                                "preconditions": [],
                                "steps": ["打开注册"],
                                "expected": "可打开",
                                "design_method": "场景法",
                                "requirement_ref": "R1",
                            },
                            {
                                "id": "INVALID",
                                "title": "坏用例",
                                "priority": "P1",
                                "type": "功能",
                                "preconditions": [],
                                "steps": ["x"],
                                "expected": "y",
                                "design_method": "场景法",
                                "requirement_ref": "R-NOT-EXIST",
                            },
                        ],
                    },
                    {"op": "validate_and_export", "fill_gaps": False},
                ],
            })

    run_chat(store, job.id, "补两条", MixedLLM())
    after = {c["id"] for c in parse_cases(store.output_dir(job.id) / "testcases.md")}
    assert "TC-REG-008" in after
    assert "INVALID" not in after
    assert before_ids <= after


def test_update_local_llm_writes_and_masks(tmp_path):
    from qagent.config import mask_api_key
    (tmp_path / "qagent.yaml").write_text("language: zh\n", encoding="utf-8")
    api_key = _dummy_api_key()
    out = update_local_llm(api_key=api_key, model="gpt-test", workspace=tmp_path)
    assert out["api_key_configured"] is True
    assert out["api_key_hint"] == mask_api_key(api_key)
    assert out["api_key_source"] == "file"
    assert api_key not in json.dumps(out)
    text = (tmp_path / "qagent.local.yaml").read_text(encoding="utf-8")
    assert api_key in text
    again = update_local_llm(base_url="https://example.com/v1/", workspace=tmp_path)
    assert again["model"] == "gpt-test"
    assert again["base_url"] == "https://example.com/v1"
    assert public_llm_settings(tmp_path)["api_key_configured"] is True


def test_http_settings_roundtrip(tmp_path, monkeypatch):
    from qagent.config import mask_api_key
    (tmp_path / "qagent.yaml").write_text("language: zh\n", encoding="utf-8")
    monkeypatch.setattr("qagent.config.find_workspace_root", lambda start=None: tmp_path)
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    httpd = ThreadingHTTPServer(("127.0.0.1", 0), create_handler(service))
    thread = threading.Thread(target=httpd.serve_forever, daemon=True)
    thread.start()
    port = httpd.server_address[1]
    try:
        empty = _http_json(port, "GET", "/api/settings")
        assert empty["api_key_configured"] is False
        api_key = _dummy_api_key()
        body = json.dumps({"api_key": api_key, "model": "demo-model"}).encode("utf-8")
        saved = _http_json(port, "POST", "/api/settings", body=body, headers={
            "Content-Type": "application/json",
        })
        assert saved["api_key_configured"] is True
        assert api_key not in json.dumps(saved)
        assert saved["model"] == "demo-model"
        got = _http_json(port, "GET", "/api/settings")
        assert got["api_key_hint"] == mask_api_key(api_key)
    finally:
        httpd.shutdown()


def test_public_job_includes_deliverables(tmp_path):
    store = JobStore(tmp_path / "jobs")
    service = QAgentService(store, llm_factory=lambda: MockLLM({}))
    job = store.create()
    (store.output_dir(job.id) / "test-requirements.md").write_text("# 需求\n", encoding="utf-8")
    (store.output_dir(job.id) / "test-requirements.drawio").write_text("<mxfile/>", encoding="utf-8")
    store.refresh_artifacts(job.id)
    got = service.get_job(job.id)
    titles = [d["title"] for d in got["deliverables"]]
    assert titles == ["测试需求", "需求导图"]
    assert got["deliverables"][0]["role"] == "测什么、不测什么"
    assert got["deliverables"][1]["file"] == "test-requirements.drawio"
