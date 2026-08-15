# /// script
# requires-python = ">=3.9"
# dependencies = ["pyyaml", "openpyxl"]
# ///
"""把 testcases.md 中的 YAML 用例导出为 Excel。

用法:
    python scripts/export_xlsx.py output/testcases.md --out output/testcases.xlsx
"""

import argparse
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from qa_common import parse_cases, parse_requirement_ids, validate

COLUMNS = [
    ("id", "ID", 14),
    ("title", "标题", 36),
    ("priority", "优先级", 8),
    ("type", "类型", 8),
    ("preconditions", "前置条件", 28),
    ("steps", "步骤", 48),
    ("expected", "预期结果", 36),
    ("design_method", "设计方法", 12),
    ("requirement_ref", "需求追溯", 12),
]


def cell_value(case: dict, field: str) -> str:
    value = case.get(field)
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(f"{i}. {item}" if field == "steps" else str(item)
                         for i, item in enumerate(value, start=1))
    return str(value)


def main() -> int:
    parser = argparse.ArgumentParser(description="导出测试用例到 xlsx")
    parser.add_argument("cases", type=Path, help="testcases.md 路径")
    parser.add_argument("--out", type=Path, required=True, help="输出 xlsx 路径")
    parser.add_argument("--plan", type=Path, default=None,
                        help="test-plan.md 路径（提供时先做校验，不通过则拒绝导出）")
    args = parser.parse_args()

    try:
        cases = parse_cases(args.cases)
    except (OSError, ValueError) as exc:
        print(f"ERROR: {exc}")
        return 1

    if args.plan is not None:
        try:
            requirement_ids = parse_requirement_ids(args.plan)
        except (OSError, ValueError) as exc:
            print(f"ERROR: {exc}")
            return 1
        errors, _ = validate(cases, requirement_ids)
        if errors:
            for error in errors:
                print(f"ERROR: {error}")
            print("FAILED: 用例未通过校验，拒绝导出。请先修正后重试，"
                  "或不带 --plan 参数强制导出。")
            return 1

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "测试用例"

    for col, (_, header, width) in enumerate(COLUMNS, start=1):
        cell = sheet.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        sheet.column_dimensions[get_column_letter(col)].width = width
    sheet.freeze_panes = "A2"

    for row, case in enumerate(cases, start=2):
        for col, (field, _, _) in enumerate(COLUMNS, start=1):
            cell = sheet.cell(row=row, column=col, value=cell_value(case, field))
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    args.out.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(args.out)
    print(f"OK: 已导出 {len(cases)} 条用例到 {args.out}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
