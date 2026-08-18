"""Excel 导出器。"""

from __future__ import annotations

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from qagent.exporters import ExportContext


def _cell_value(case: dict, field: str) -> str:
    value = case.get(field)
    if value is None:
        return ""
    if isinstance(value, list):
        return "\n".join(
            f"{i}. {item}" if field == "steps" else str(item)
            for i, item in enumerate(value, start=1)
        )
    return str(value)


class XlsxExporter:
    def export(self, ctx: ExportContext) -> Path:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "测试用例"

        columns = ctx.schema.export_columns
        for col, (field, header, width) in enumerate(columns, start=1):
            cell = sheet.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            sheet.column_dimensions[get_column_letter(col)].width = width
        sheet.freeze_panes = "A2"

        for row, case in enumerate(ctx.cases, start=2):
            for col, (field, _, _) in enumerate(columns, start=1):
                cell = sheet.cell(row=row, column=col, value=_cell_value(case, field))
                cell.alignment = Alignment(wrap_text=True, vertical="top")

        ctx.output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(ctx.output_path)
        return ctx.output_path
